import datetime
import re
from enum import Enum

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from config import Config
from utils import stemmer, tokenize
from copy import copy


COLUMNS_TO_READ_FROM_TOLOKA = ['B', 'D', 'E', '|', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W',
                               'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL']

COLUMNS_TO_READ_FROM_TEMPLATE = ['A', 'B', 'C', 'D']


def chunks(l, n):
    n = max(1, n)
    return [l[i:i+n] for i in range(0, len(l), n)]


def split_workbook(wb: Workbook):
    result = {}
    for chunk in chunks(wb.sheetnames, Config.SHEETS_IN_FILE):
        filename = chunk[0][0] + '-' + chunk[-1][0] + '_' + datetime.datetime.now().strftime('%d-%m-%YT%H:%M:%S')
        output_workbook = create_workbook()
        for sheet in wb.sheetnames:
            work_sheet = output_workbook.create_sheet(sheet)
            for row_data in wb.get_sheet_by_name(sheet).iter_rows():
                for row_cell in row_data:
                    work_sheet[row_cell.coordinate].value = row_cell.value
                    if row_cell.has_style:
                        work_sheet[row_cell.coordinate].fill = copy(row_cell.fill)
                        work_sheet[row_cell.coordinate].font = copy(row_cell.font)

        result[filename] = output_workbook
    return result

class Colors(Enum):
    WHITE = 'FFFFFF'
    GREY = 'cccccc'
    ORANGE = 'f9cb9c'
    TURQUOISE = '40E0D0'
    GREEN = 'b6d7a8'
    RED = 'FF0000'
    BLUE = '6666FF'


def create_workbook():
    workbook = Workbook(iso_dates=True)
    workbook.remove(workbook["Sheet"])
    return workbook


def as_text(value):
    if value is None:
        return ""
    return str(value)


def resize_columns(sheet):
    for column_cells in sheet.columns:
        length = max([len(as_text(cell.value)) for cell in column_cells]) + 2
        sheet.column_dimensions[chr(ord('A') - 1 + column_cells[0].column)].width = float(length)


def set_left_aligment(sheet):
    for column_cells in sheet.columns:
        for cell in column_cells:
            cell.alignment = Alignment(horizontal='left')


def check_toloka(workbook):
    toloka_workbook = load_workbook(Config.TOLOKA_PATH)
    toloka_sheet = toloka_workbook.get_sheet_by_name(toloka_workbook.sheetnames[0])

    template_workbook = load_workbook(Config.WORD_TEMOLATE_PATH)
    template_sheet = template_workbook.get_sheet_by_name(template_workbook.sheetnames[0])

    toloka_raw = 2

    last_raw = {}

    while True:
        if not toloka_sheet[f'D{toloka_raw}'].value:
            break
        sheet_name = toloka_sheet[f'D{toloka_raw}'].value.lower()
        if ' '.join(stemmer(sheet_name.split(' '))) not in [' '.join(stemmer(sheetname.split(' '))) for sheetname in workbook.sheetnames]:
            work_sheet = workbook.create_sheet(sheet_name)
            last_raw[sheet_name] = 1
            work_raw = 1
        else:
            work_sheet = workbook.get_sheet_by_name(sheet_name)
            work_raw = last_raw[sheet_name]
            for column in range(0, 6):
                work_sheet[f'{chr(ord("A") + column)}{work_raw - 1}'].fill = PatternFill("solid",
                                                                                         fgColor=Colors.GREEN.value)

        work_raw_from_toloka = work_raw
        unique_stems = []
        unique_words_counter = 0
        for column in COLUMNS_TO_READ_FROM_TOLOKA:
            if column == '|':
                work_raw_from_toloka += 1
                continue
            elif column == 'B':
                work_sheet[f'G{work_raw}'] = toloka_sheet[f'{column}{toloka_raw}'].value
                work_sheet[f'G{work_raw}'].font = Font(bold=True)
                work_sheet[f'G{work_raw}'].fill = PatternFill("solid", fgColor=Colors.ORANGE.value)
                continue
            elif column == 'D':
                work_sheet[f'F{work_raw_from_toloka}'].font = Font(bold=True)

            toloka_value = ' '.join(str(toloka_sheet[f'{column}{toloka_raw}'].value).capitalize().split()).replace('.0',
                                                                                                                   '')
            toloka_value = re.sub(r'(,)(\S)', r', \2', toloka_value)

            tokens = tokenize(toloka_value)
            stems = stemmer(tokens)
            for stem in stems:
                if stem not in unique_stems:
                    unique_stems.append(stem)
                    unique_words_counter += 1

            work_sheet[f'F{work_raw_from_toloka}'] = toloka_value
            work_raw_from_toloka += 1

        work_sheet[f'E{work_raw}'] = 'Ядро'
        work_sheet[f'E{work_raw}'].font = Font(bold=True)
        work_sheet[f'E{work_raw + 1}'] = unique_words_counter

        max_work_raw_from_template = 0

        for column in COLUMNS_TO_READ_FROM_TEMPLATE:
            raw_template = 1
            work_raw_from_template = work_raw
            while template_sheet[f'{column}{raw_template}'].value:
                template_value = template_sheet[f'{column}{raw_template}'].value
                if work_raw_from_template > work_raw:
                    if any([stemmer([token])[0] in unique_stems for token in tokenize(template_value)]):
                        work_sheet[f'{column}{work_raw_from_template}'].fill = PatternFill("solid",
                                                                                           fgColor=Colors.GREY.value)
                else:
                    work_sheet[f'{column}{work_raw_from_template}'].font = Font(bold=True)
                    work_sheet[f'{column}{work_raw_from_template}'].fill = PatternFill("solid",
                                                                                       fgColor=Colors.ORANGE.value)
                work_sheet[f'{column}{work_raw_from_template}'] = template_value
                work_raw_from_template += 1
                raw_template += 1

            max_work_raw_from_template = work_raw_from_template if work_raw_from_template > max_work_raw_from_template \
                else max_work_raw_from_template

        work_raw = work_raw_from_toloka if work_raw_from_toloka > max_work_raw_from_template else max_work_raw_from_template

        toloka_raw += 1
        work_raw += 1

        last_raw[sheet_name] = work_raw

    for sheet in workbook.sheetnames:
        resize_columns(sheet=workbook.get_sheet_by_name(sheet))
        set_left_aligment(sheet=workbook.get_sheet_by_name(sheet))

    workbook._sheets.sort(key=lambda ws: ws.title)

    return workbook
